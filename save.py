import os

from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook


class Save:
    filename = "data.xlsx"

    def __init__(self):
        # Vérifier si le fichier existe
        if os.path.exists(self.filename):
            self.wb = load_workbook(self.filename)
            self.ws = self.wb.active
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.append([
                "Nom", "Boîte Postale", "Téléphones", "WhatsApp", "Fax", "Email", "Site Web",
                "Secteur d'activité", "Localisation"
            ])

    def add(self, name, bp, tel, whatsapp, fax, email, website, sector, location):
        self.ws.append([name, bp, tel, whatsapp, fax, email, website, sector, location])

    def close(self):
        self.wb.save(self.filename)
