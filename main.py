from playwright.sync_api import sync_playwright
import urllib.parse
import string
import sqlite3
from openpyxl import Workbook
import re

BASE_URL = "https://www.doualazoom.com"

# Initialisation du fichier Excel
wb = Workbook()
ws = wb.active
ws.append(["Nom", "Boîte Postale", "Téléphones", "WhatsApp", "Fax", "Email", "Site Web", "Secteur d'activité", "Localisation"])

# Initialisation de la base de données SQLite
conn = sqlite3.connect("entreprises.db")
cursor = conn.cursor()
cursor.execute("""
    CREATE TABLE IF NOT EXISTS entreprises (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nom TEXT,
        boite_postale TEXT,
        telephones TEXT,
        whatsapp TEXT,
        fax TEXT,
        email TEXT,
        site_web TEXT,
        secteur_activite TEXT,
        localisation TEXT
    )
""")
conn.commit()

# Fonction pour normaliser une URL
def normalize_url(base_url, relative_url):
    if not relative_url:
        return None
    decoded_url = relative_url.replace("httpts://", "https://")  # Corriger les erreurs d'URL
    full_url = f"{base_url}{decoded_url}" if relative_url.startswith("/") else relative_url
    return full_url


# Fonction pour extraire le nom de l'entreprise depuis l'URL
def extract_name_from_url(url):
    try:
        # Extraire la partie contenant le nom
        nom_partie_encodee = url.split("/fr/activite/alpha/A/")[-1].split("/")[0]
        # Décoder la partie encodée
        nom_entreprise = urllib.parse.unquote(nom_partie_encodee).strip()
        return nom_entreprise
    except Exception as e:
        print(f"⚠️ Erreur lors de l'extraction du nom depuis l'URL : {e}")
        return "Non renseigné"

# Fonction pour extraire les détails d'une entreprise
def extract_entreprise_details(page):
    try:
        print(f"\n⏳ Extraction des détails de l'entreprise sur {page.url}...")
        
        # Extraire le nom de l'entreprise depuis l'URL
        name = extract_name_from_url(page.url)
        print('----------------ijrijer-', name)
        
        # Boîte postale
        bp_text = page.evaluate(
            """
            () => {
                const el = Array.from(document.querySelectorAll('div')).find(div => div.textContent.includes('Boite postale'));
                return el ? el.querySelector('a')?.textContent.trim() : null;
            }
            """
        )
        postal_code = bp_text if bp_text else "Non renseigné"
        
        # Téléphones, Fax, WhatsApp
        phone_elements = page.query_selector_all("div:has(b)")
        phones = {"Téléphones": [], "WhatsApp": [], "Fax": []}
        seen_numbers = set()
        for el in phone_elements:
            label = el.query_selector("b").text_content().strip() if el.query_selector("b") else ""
            value = el.query_selector("a").text_content().strip() if el.query_selector("a") else ""
            if value and value not in seen_numbers:
                seen_numbers.add(value)
                if "Téléphone" in label:
                    phones["Téléphones"].append(value)
                elif "Fax" in label or "Viber" in label:
                    phones["Fax"].append(value)
                elif "WhatsApp" in label:
                    phones["WhatsApp"].append(value)
        
        # Email
        email_el = page.query_selector("a[href^='mailto']")
        email = email_el.text_content().strip() if email_el else "Non renseigné"
        
        # Site web
        website_el = page.query_selector("a[target='_blanc']")
        website = normalize_url(BASE_URL, website_el.get_attribute("href")) if website_el else "Non renseigné"
        
        # Secteur d'activité
        secteur_activite = page.evaluate(
            """
            () => {
                const bElement = Array.from(document.querySelectorAll('b')).find(b => b.textContent.trim() === 'Secteur d\\'activité:');
                return bElement ? bElement.nextElementSibling?.textContent.trim() : 'Non renseigné';
            }
            """
        )
        
        # Localisation
        location_text = page.text_content(".div_titre_surlacarte") if page.query_selector(".div_titre_surlacarte") else ""
        location_match = re.search(r"latitude ([\d.]+),.*longitude ([\d.]+)", location_text)
        location = (location_match.group(1), location_match.group(2)) if location_match else ("Non renseigné", "Non renseigné")
        
        # Préparer les données extraites
        extracted_data = {
            "Nom": name,
            "Boîte Postale": postal_code,
            "Téléphones": ", ".join(phones["Téléphones"]),
            "WhatsApp": ", ".join(phones["WhatsApp"]),
            "Fax": ", ".join(phones["Fax"]),
            "Email": email,
            "Site Web": website,
            "Secteur d'activité": secteur_activite,
            "Localisation": f"({location[0]}, {location[1]})"
        }
        print(f"✅ Données extraites : {extracted_data}")
        
        # Enregistrer dans la base de données SQLite
        cursor.execute("""
            INSERT INTO entreprises (nom, boite_postale, telephones, whatsapp, fax, email, site_web, secteur_activite, localisation)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            extracted_data["Nom"],
            extracted_data["Boîte Postale"],
            extracted_data["Téléphones"],
            extracted_data["WhatsApp"],
            extracted_data["Fax"],
            extracted_data["Email"],
            extracted_data["Site Web"],
            extracted_data["Secteur d'activité"],
            extracted_data["Localisation"]
        ))
        conn.commit()
        
        # Enregistrer dans le fichier Excel
        ws.append([
            extracted_data["Nom"],
            extracted_data["Boîte Postale"],
            extracted_data["Téléphones"],
            extracted_data["WhatsApp"],
            extracted_data["Fax"],
            extracted_data["Email"],
            extracted_data["Site Web"],
            extracted_data["Secteur d'activité"],
            extracted_data["Localisation"]
        ])
        wb.save("entreprises.xlsx")  # Sauvegarder immédiatement
        
        return extracted_data
    except Exception as e:
        print(f"⚠️ Erreur lors de l'extraction des détails de l'entreprise sur {page.url}: {e}")
        return None

# Fonction pour extraire toutes les entreprises d'une page
def extract_all_companies_on_page(context, url):
    try:
        print(f"\n🔄 Chargement de la liste des entreprises depuis {url}...")
        page = context.new_page()
        page.goto(url, timeout=60000)
        page.wait_for_selector(".div_list_nomentreprise a", timeout=15000)
        
        # Récupérer tous les liens des entreprises
        entreprise_links = page.query_selector_all(".div_list_nomentreprise a")
        if len(entreprise_links) == 0:
            print("⚠️ Aucune entreprise trouvée sur cette page.")
            page.close()
            return False
        
        # Traiter au maximum 25 entreprises par page
        processed_count = 0
        for i, link in enumerate(entreprise_links[:25]):
            raw_url = link.get_attribute("href")
            clean_url = normalize_url(BASE_URL, raw_url)
            
            if not clean_url:
                print(f"⚠️ URL invalide pour l'entreprise {i + 1}, ignorée.")
                continue
            
            # Ouvrir une nouvelle page pour l'entreprise
            new_page = context.new_page()
            print(f"🔗 Ouvrir l'onglet pour {clean_url}")
            
            try:
                new_page.goto(clean_url, timeout=60000)
                new_page.wait_for_load_state("load", timeout=30000)
                
                # Extraire les détails de l'entreprise
                extract_entreprise_details(new_page)
                
                # Fermer l'onglet après l'extraction
                print(f"🔒 Fermeture de l'onglet pour l'entreprise {i + 1}.")
                new_page.close()
                
                processed_count += 1
            except Exception as e:
                print(f"⚠️ Erreur lors de l'ouverture de l'entreprise {i + 1}: {e}")
                new_page.close()
        
        page.close()  # Fermer la page principale
        return processed_count >= 25  # Retourne True si 25 entreprises ont été traitées
    except Exception as e:
        print(f"⚠️ Erreur lors du traitement des entreprises sur {url}: {e}")
        return False

# Fonction principale
def main():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context()
        
        # Itérer sur toutes les lettres de l'alphabet
        for letter in string.ascii_uppercase:  # De 'A' à 'Z'
            print(f"\n🔍 Démarrage de l'extraction pour la lettre {letter}...")
            page_number = 1
            
            while True:
                url = f"{BASE_URL}/fr/activite/alpha/{letter}?page={page_number}"
                print(f"\n🔄 Extraction des données pour la lettre {letter}, page {page_number}...")
                
                # Extraire les entreprises de la page actuelle
                should_continue = extract_all_companies_on_page(context, url)
                
                # Vérifier s'il y a une page suivante
                if not should_continue:
                    print(f"✅ Fin de l'extraction pour la lettre {letter}.")
                    break
                
                page_number += 1
        
        browser.close()
        conn.close()  # Fermer la connexion à la base de données

# Lancer le script
if __name__ == "__main__":
    main()





























"""from playwright.sync_api import sync_playwright
import re
import urllib.parse

BASE_URL = "https://www.doualazoom.com"

# Fonction pour normaliser une URL
def normalize_url(base_url, relative_url):
    if not relative_url:
        return None
    decoded_url = relative_url.replace("httpts://", "https://")  # Corriger les erreurs d'URL
    full_url = f"{base_url}{decoded_url}" if relative_url.startswith("/") else relative_url
    return full_url

# Fonction pour extraire le nom de l'entreprise depuis l'URL
def extract_name_from_url(url):
    try:
        # Extraire la partie contenant le nom
        nom_partie_encodee = url.split("/fr/activite/alpha/A/")[-1].split("/")[0]
        # Décoder la partie encodée
        nom_entreprise = urllib.parse.unquote(nom_partie_encodee).strip()
        return nom_entreprise
    except Exception as e:
        print(f"⚠️ Erreur lors de l'extraction du nom depuis l'URL : {e}")
        return "Non renseigné"

# Fonction pour extraire les détails d'une entreprise
def extract_entreprise_details(page):
    try:
        print(f"\n⏳ Extraction des détails de l'entreprise sur {page.url}...")
        
        # Extraire le nom de l'entreprise
        name = extract_name_from_url(page.url)
        print('jhksfshfssssss', name)
            
        # Extraire la boîte postale
        bp_text = page.evaluate(
            ---
            () => {
                const el = Array.from(document.querySelectorAll('div')).find(div => div.textContent.includes('Boite postale'));
                return el ? el.querySelector('a')?.textContent.trim() : null;
            }
            ---
        )
        postal_code = bp_text if bp_text else "Non renseigné"
        
        # Extraire les numéros de téléphone, fax et WhatsApp
        phone_elements = page.query_selector_all("div:has(b)")
        phones = {"Téléphones": [], "WhatsApp": [], "Fax": []}
        seen_numbers = set()  # Pour éviter les doublons
        for el in phone_elements:
            label = el.query_selector("b").text_content().strip() if el.query_selector("b") else ""
            value = el.query_selector("a").text_content().strip() if el.query_selector("a") else ""
            if value and value not in seen_numbers:
                seen_numbers.add(value)
                if "Téléphone" in label:
                    phones["Téléphones"].append(value)
                elif "Fax" in label or "Viber" in label:
                    phones["Fax"].append(value)
                elif "WhatsApp" in label:
                    phones["WhatsApp"].append(value)
        
        # Extraire l'email
        email_el = page.query_selector("a[href^='mailto']")
        email = email_el.text_content().strip() if email_el else "Non renseigné"
        
        # Extraire le site web
        website_el = page.query_selector("a[target='_blanc']")
        website = normalize_url(BASE_URL, website_el.get_attribute("href")) if website_el else "Non renseigné"
        
        # Extraire le secteur d'activité
        secteur_activite = page.evaluate(
            
            () => {
                const bElement = Array.from(document.querySelectorAll('b')).find(b => b.textContent.trim() === 'Secteur d\\'activité:');
                return bElement ? bElement.nextElementSibling?.textContent.trim() : 'Non renseigné';
            }
            
        )
        
        # Extraire la localisation
        location_text = page.text_content(".div_titre_surlacarte") if page.query_selector(".div_titre_surlacarte") else ""
        location_match = re.search(r"latitude ([\d.]+),.*longitude ([\d.]+)", location_text)
        location = (location_match.group(1), location_match.group(2)) if location_match else ("Non renseigné", "Non renseigné")
        
        # Afficher les détails récupérés
        extracted_data = {
            "Nom": name,
            "Boîte Postale": postal_code,
            "Téléphones": phones,
            "Email": email,
            "Site Web": website,
            "Secteur d'activité": secteur_activite,
            "Localisation": location,
        }
        print(f"✅ Données extraites : {extracted_data}")
        return extracted_data
    except Exception as e:
        print(f"⚠️ Erreur lors de l'extraction des détails de l'entreprise sur {page.url}: {e}")
        return None

# Fonction principale
def main():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context()
        letter = 'A'
        page_number = 1
        
        while True:
            url = f"{BASE_URL}/fr/activite/alpha/{letter}?page={page_number}"
            print(f"\n🔍 Extraction des données depuis {url}...")
            
            # Charger la liste des entreprises
            page = context.new_page()
            page.goto(url, timeout=60000)
            page.wait_for_selector(".div_list_nomentreprise a", timeout=15000)
            
            # Récupérer tous les liens des entreprises
            entreprise_links = page.query_selector_all(".div_list_nomentreprise a")
            if len(entreprise_links) == 0:
                print("⚠️ Aucune entreprise trouvée sur cette page.")
                break
            
            # Traiter chaque lien d'entreprise individuellement
            for i, link in enumerate(entreprise_links[:25]):  # Limiter à 25 entreprises
                raw_url = link.get_attribute("href")
                clean_url = normalize_url(BASE_URL, raw_url)
                
                if not clean_url:
                    print(f"⚠️ URL invalide pour l'entreprise {i + 1}, ignorée.")
                    continue
                
                # Ouvrir une nouvelle page pour l'entreprise
                new_page = context.new_page()
                print(f"🔗 Ouvrir l'onglet pour {clean_url}")
                
                try:
                    new_page.goto(clean_url, timeout=60000)
                    new_page.wait_for_load_state("load", timeout=30000)
                    
                    # Extraire les détails de l'entreprise
                    extract_entreprise_details(new_page)
                    
                    # Fermer l'onglet après l'extraction
                    print(f"🔒 Fermeture de l'onglet pour l'entreprise {i + 1}.")
                    new_page.close()
                except Exception as e:
                    print(f"⚠️ Erreur lors de l'ouverture de l'entreprise {i + 1}: {e}")
                    new_page.close()
            
            page.close()  # Fermer la page principale
            page_number += 1
            
            # Limiter à deux pages pour éviter un scraping infini
            if page_number > 2:
                break
        
        browser.close()

# Lancer le script
if __name__ == "__main__":
    main()"""
















"""from playwright.sync_api import sync_playwright
import pandas as pd
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import os
import sqlite3
import requests

# URL de base
BASE_URL = "https://www.doualazoom.com"

# Fichiers de sortie
EXCEL_FILE = "entreprises.xlsx"
PDF_FOLDER = "pdfs"
IMAGES_FOLDER = "images"
SQLITE_DB = "entreprises.db"

# Créer les dossiers s'ils n'existent pas
os.makedirs(PDF_FOLDER, exist_ok=True)
os.makedirs(IMAGES_FOLDER, exist_ok=True)

# Initialisation des données
data = []

# Connexion à la base de données SQLite
def init_db():
    Initialise la base de données SQLite.
    conn = sqlite3.connect(SQLITE_DB)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS entreprises (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nom TEXT,
            telephone TEXT,
            email TEXT,
            adresse TEXT,
            image_path TEXT
        )
    ''')
    conn.commit()
    return conn

def save_to_db(conn, nom, telephone, email, adresse, image_path):
    Enregistre une entreprise dans la base de données SQLite.
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO entreprises (nom, telephone, email, adresse, image_path)
        VALUES (?, ?, ?, ?, ?)
    ''', (nom, telephone, email, adresse, image_path))
    conn.commit()

def download_image(url, save_path):
    Télécharge une image depuis une URL et l'enregistre localement.
    try:
        response = requests.get(url, stream=True)
        if response.status_code == 200:
            with open(save_path, 'wb') as file:
                for chunk in response.iter_content(1024):
                    file.write(chunk)
            print(f"✅ Image téléchargée : {save_path}")
        else:
            print(f"⚠️ Impossible de télécharger l'image depuis {url}")
    except Exception as e:
        print(f"⚠️ Erreur lors du téléchargement de l'image : {e}")

def save_to_excel():
    Enregistre les données dans un fichier Excel.
    df = pd.DataFrame(data, columns=["Nom", "Téléphone", "Email", "Adresse", "Image"])
    df.to_excel(EXCEL_FILE, index=False, engine='openpyxl')
    print(f"📊 Données sauvegardées dans {EXCEL_FILE}")


def save_allto_pdf(nom, telephone, email, adresse, image_path):
    Génère un fichier PDF pour une entreprise.
    pdf_path = os.path.join(PDF_FOLDER, f"{nom.replace('/', '_')}.pdf")
    c = canvas.Canvas(pdf_path, pagesize=letter)
    width, height = letter
    y = height - 40  # Position de départ

    c.setFont("Helvetica", 12)
    c.drawString(30, y, f"📌 Nom: {nom}")
    y -= 20
    c.drawString(30, y, f"📞 Téléphone: {telephone}")
    y -= 20
    c.drawString(30, y, f"📧 Email: {email}")
    y -= 20
    c.drawString(30, y, f"📍 Adresse: {adresse}")
    y -= 20

    # Vérifier si l'image est valide avant de l'ajouter
    if image_path and os.path.exists(image_path):
        try:
            c.drawImage(image_path, 30, y - 100, width=100, height=100)
            y -= 120  # Ajustement de la position après l'ajout de l'image
        except Exception as e:
            print(f"⚠️ Erreur lors de l'ajout de l'image dans le PDF : {e}")

    c.save()
    print(f"📄 Fichier PDF généré : {pdf_path}")

def save_to_pdf():
    Génère un fichier PDF unique contenant toutes les entreprises.
    pdf_path = os.path.join(PDF_FOLDER, "entreprises_doul.pdf")
    c = canvas.Canvas(pdf_path, pagesize=letter)
    width, height = letter
    y = height - 40  # Position de départ

    c.setFont("Helvetica", 12)
    for entry in data:
        nom, telephone, email, adresse, image_path = entry
        c.drawString(30, y, f"📌 Nom: {nom}")
        y -= 20
        c.drawString(30, y, f"📞 Téléphone: {telephone}")
        y -= 20
        c.drawString(30, y, f"📧 Email: {email}")
        y -= 20
        c.drawString(30, y, f"📍 Adresse: {adresse}")
        y -= 40  # Saut de ligne après chaque entreprise

        if image_path and os.path.exists(image_path):
            c.drawImage(image_path, 30, y - 100, width=100, height=100)
            y -= 120  # Ajuster la hauteur après l'image
        
        if y < 100:  # Si on est proche du bas de page, on crée une nouvelle page
            c.showPage()
            y = height - 40
    
    c.save()
    print(f"📄 Fichier PDF unique généré : {pdf_path}")
    


def extract_page_data(page, conn):
    Extrait les données d'une page d'entreprises et les enregistre.
    page.wait_for_selector('.div_categorielist_detail02', timeout=10000)

    entreprises = page.query_selector_all('.div_categorielist_detail02')
    if not entreprises:
        print("⚠️ Aucune entreprise trouvée sur cette page.")
        return False

    for entreprise in entreprises:
        nom_element = entreprise.query_selector('.div_list_nomentreprise a')
        nom = nom_element.inner_text().strip() if nom_element else "Nom inconnu"

        details_element = entreprise.query_selector('.div_list_detailentreprise')
        details = details_element.inner_text().strip() if details_element else "Aucun détail"

        telephone = None
        if "Téléphone:" in details:
            telephone = details.split("Téléphone:")[1].split("\n")[0].strip()

        email_element = entreprise.query_selector('.div_list_detailentreprise a[href^="mailto:"]')
        email = email_element.get_attribute('href').replace("mailto:", "").strip() if email_element else "Aucun email"

        adresse_parts = []
        adresse_elements = entreprise.query_selector_all('.div_list_detailentreprise a')
        for element in adresse_elements:
            href = element.get_attribute('href')
            if href and ("ville" in href or "alpha" in href):
                adresse_parts.append(element.inner_text().strip())
        adresse = ", ".join(adresse_parts) if adresse_parts else "Aucune adresse"

        # Image
         #img_element = entreprise.query_selector('.div_smalllogo img')

        img_element = page.query_selector('.div_smalllogo img')
        if img_element:
            img_url = img_element.get_attribute('src')
            full_img_url = BASE_URL + img_url if img_url.startswith("/") else img_url
            image_name = f"{nom.replace('/', '_')}.jpg"  # Nom du fichier image
            image_path = os.path.join(IMAGES_FOLDER, image_name)
            download_image(full_img_url, image_path)

        # Ajouter aux données
        data.append([nom, telephone, email, adresse, image_path])
        print(f"📌 {nom} ajouté aux fichiers.")

        # Sauvegarde dans la base de données
        save_to_db(conn, nom, telephone, email, adresse, image_path)

        # Générer un PDF pour cette entreprise
        save_to_pdf()

        # Sauvegarder les données dans un fichier Excel
        save_to_excel()


    # Vérifier s'il y a une page suivante
    next_button = page.query_selector('.pagination .next a')
    return next_button is not None

def extract_all_data():
    Parcourt toutes les lettres et toutes les pages pour extraire les entreprises.
    conn = init_db()  # Initialiser la base de données
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()

        for letter in 'ABCDEFGHIJKLMNOPQRSTUVWXYZ':
            page_number = 1
            while True:
                url = f"{BASE_URL}/fr/activite/alpha/{letter}?page={page_number}"
                print(f"🔍 Extraction des données depuis {url}...")

                page.goto(url, timeout=1800000)
                has_next_page = extract_page_data(page, conn)

                if not has_next_page:
                    print(f"✅ Fin de l'extraction pour la lettre {letter}.")
                    break

                page_number += 1

        browser.close()
    conn.close()  # Fermer la connexion à la base de données

# Lancer l'extraction
extract_all_data()
print("✅ Extraction terminée !")"""
























"""from playwright.sync_api import sync_playwright

BASE_URL = "https://www.doualazoom.com"

def extract_page_data(page):
    Extrait les données d'une page d'entreprises et retourne si une page suivante existe. 
    
    # Attendre le chargement des entreprises
    page.wait_for_selector('.div_categorielist_detail02', timeout=10000)  # Timeout de 10s

    entreprises = page.query_selector_all('.div_categorielist_detail02')
    
    if not entreprises:
        print("⚠️ Aucune entreprise trouvée sur cette page.")
        return False  # Arrêter si aucune entreprise

    for entreprise in entreprises:
        # Nom de l'entreprise
        nom_element = entreprise.query_selector('.div_list_nomentreprise a')
        nom = nom_element.inner_text().strip() if nom_element else "Nom inconnu"
        
        # Détails (Téléphone, Email)
        details_element = entreprise.query_selector('.div_list_detailentreprise')
        details = details_element.inner_text().strip() if details_element else "Aucun détail"
        
        # Extraction Téléphone
        telephone = None
        if "Téléphone:" in details:
            telephone = details.split("Téléphone:")[1].split("\n")[0].strip()
        
        # Extraction Email
        email_element = entreprise.query_selector('.div_list_detailentreprise a[href^="mailto:"]')
        email = email_element.get_attribute('href').replace("mailto:", "").strip() if email_element else "Aucun email"

        # Adresse (Extrait les liens pertinents)
        adresse_parts = []
        adresse_elements = entreprise.query_selector_all('.div_list_detailentreprise a')
        for element in adresse_elements:
            href = element.get_attribute('href')
            if href and ("ville" in href or "alpha" in href):
                adresse_parts.append(element.inner_text().strip())
        adresse = ", ".join(adresse_parts) if adresse_parts else "Aucune adresse"

        # Image
        img_element = page.query_selector('.div_smalllogo img')
        if img_element:
            img_url = img_element.get_attribute('src')
            full_img_url = "https://www.doualazoom.com" + img_url
            print(f"URL de l'image : {full_img_url}")
                
            # Télécharger l'image
            #file_name = "image_downloaded.gif"  # ou tout autre nom souhaité
            #download_image(full_img_url, file_name)
        else:
            print("Aucune image trouvée.")

        # Affichage des résultats
        print(f"📌 Nom: {nom}")
        print(f"📞 Téléphone: {telephone}")
        print(f"📧 Email: {email}")
        print(f"📍 Adresse: {adresse}")
        print(f"🖼️ Image: {full_img_url}")
        print("-" * 50)

    # Vérifier s'il y a une page suivante
    next_button = page.query_selector('.pagination .next a')
    return next_button is not None  # Retourne True si une autre page existe

def extract_all_data():
    Parcourt toutes les lettres et toutes les pages pour extraire les entreprises. 
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)  # headless=False pour voir le navigateur
        page = browser.new_page()

        for letter in 'ABCDEFGHIJKLMNOPQRSTUVWXYZ':
            page_number = 1
            while True:
                url = f"{BASE_URL}/fr/activite/alpha/{letter}?page={page_number}"
                print(f"🔍 Extraction des données depuis {url}...")
                
                page.goto(url, timeout=80000)
                has_next_page = extract_page_data(page)

                if not has_next_page:
                    print(f"✅ Fin de l'extraction pour la lettre {letter}.")
                    break  # Passer à la lettre suivante

                page_number += 1

        browser.close()

# Lancer l'extraction
extract_all_data()
print("✅ Extraction terminée !")"""































"""import requests
from playwright.sync_api import sync_playwright

def download_image(image_url, file_name):
    # Télécharger l'image
    response = requests.get(image_url)
    
    if response.status_code == 200:
        with open(file_name, 'wb') as file:
            file.write(response.content)
        print(f"L'image a été téléchargée avec succès sous le nom {file_name}")
    else:
        print("Erreur lors du téléchargement de l'image.")

def scrape_and_download_image(url):
    with sync_playwright() as p:
        browser = p.chromium.launch()
        page = browser.new_page()
        page.goto(url)

        # Attendre que l'image soit visible (en fonction du sélecteur)
        page.wait_for_selector('.div_smalllogo img')

        # Extraire l'URL de l'image
        img_element = page.query_selector('.div_smalllogo img')
        if img_element:
            img_url = img_element.get_attribute('src')
            full_img_url = "https://www.doualazoom.com" + img_url
            print(f"URL de l'image : {full_img_url}")
            
            # Télécharger l'image
            file_name = "image_downloaded.gif"  # ou tout autre nom souhaité
            download_image(full_img_url, file_name)
        else:
            print("Aucune image trouvée.")
        
        browser.close()

# Exemple d'URL de la page que tu veux scraper
scrape_and_download_image("https://www.doualazoom.com/fr/Index/alpha/A")"""














"""from playwright.sync_api import sync_playwright

BASE_URL = "https://www.doualazoom.com"  # Base pour compléter les chemins relatifs

def scrape_entreprises(url):
    with sync_playwright() as p:
        # Lancer le navigateur
        browser = p.chromium.launch(headless=False)  # headless=False pour voir le navigateur
        page = browser.new_page()
        
        # Accéder à la page
        page.goto(url, timeout=80000)  # Timeout à 60 secondes

        
        # Attendre que les éléments soient chargés
        page.wait_for_selector('.div_categorielist_detail02')
        
        # Récupérer les éléments des entreprises
        entreprises = page.query_selector_all('.div_categorielist_detail02')
        
        for entreprise in entreprises:
            # Récupérer le nom de l'entreprise
            nom_element = entreprise.query_selector('.div_list_nomentreprise a')
            nom = nom_element.inner_text().strip() if nom_element else "Nom inconnu"
            
            # Récupérer les détails de l'entreprise
            details_element = entreprise.query_selector('.div_list_detailentreprise')
            details = details_element.inner_text().strip() if details_element else "Aucun détail"

            # Récupérer le téléphone
            telephone = None
            if "Téléphone:" in details:
                telephone = details.split("Téléphone:")[1].split("\n")[0].strip()

            # Récupérer l'email correctement
            email_element = entreprise.query_selector('.div_list_detailentreprise a[href^="mailto:"]')
            email = email_element.get_attribute('href').replace("mailto:", "").strip() if email_element else "Aucun email"

             # Debugging: Afficher le HTML de la div contenant l'image
            logo_container = entreprise.query_selector('.div_smalllogo')
            print("HTML LOGO:", logo_container.inner_html() if logo_container else "Aucun logo trouvé")


            # Récupérer l'image et compléter le lien si nécessaire

            # Attendre que l'image soit visible (en fonction du sélecteur)
            page.wait_for_selector('.div_smalllogo img')

            # Extraire l'URL de l'image
            img_element = page.query_selector('.div_smalllogo img')
            if img_element:
                img_url = img_element.get_attribute('src')
                full_img_url = "https://www.doualazoom.com" + img_url
                print(f"URL de l'image : {full_img_url}")
                
                # Télécharger l'image
                #file_name = "image_downloaded.gif"  # ou tout autre nom souhaité
                #download_image(full_img_url, file_name)
            else:
                print("Aucune image trouvée.")

           

            # Afficher les informations extraites
            print(f"Nom: {nom}")
            print(f"Téléphone: {telephone}")
            print(f"Email: {email}")
            print(f"Image: {full_img_url}")
            print(f"Détails: {details}")
            print("-" * 40)
        
        # Fermer le navigateur
        browser.close()

# URL de la première page de la lettre "A"
url = "https://www.doualazoom.com/fr/Index/alpha/A"

# Lancer le scraper
scrape_entreprises(url)"""
































"""import requests
from bs4 import BeautifulSoup
import pandas as pd
import os
from fpdf import FPDF  # Bibliothèque pour générer des PDFs

# Fonction pour extraire les informations d'une page
def extract_page_data(url):
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')
    
    entreprises = []
    
    # Trouver tous les éléments de la liste des entreprises
    for entreprise in soup.select('.div_categorielist_title_container'):
        nom_element = entreprise.select_one('.div_list_nomentreprise a')
        details_element = entreprise.select_one('.div_list_detailentreprise')
        
        if nom_element and details_element:
            nom = nom_element.text.strip()
            
            # Extraire le téléphone
            telephone_element = details_element.find("span", dir="ltr")
            telephone = telephone_element.text.strip() if telephone_element else ""
            
            # Extraire l'email
            email_element = details_element.select_one('a[href^="mailto:"]')
            email = email_element['href'].replace('mailto:', '').strip() if email_element else ""
            
            # Extraire l'adresse
            adresse_parts = [part.text.strip() for part in details_element.select('a') if "ville" in part["href"] or "alpha" in part["href"]]
            adresse = ", ".join(adresse_parts) if adresse_parts else ""
            
            entreprises.append({
                'Nom': nom,
                'Téléphone': telephone,
                'Email': email,
                'Adresse': adresse
            })
    
    return entreprises

# Fonction pour mettre à jour le fichier CSV, Excel et PDF
def update_files(new_data, csv_filename='entreprises_doualzoom.csv', excel_filename='entreprises_doualzoom.xlsx', pdf_filename='entreprises_doualzoom.pdf'):
    # Convertir les nouvelles données en DataFrame
    new_df = pd.DataFrame(new_data)
    
    # Vérifier si le fichier CSV existe déjà
    if os.path.exists(csv_filename):
        # Charger les données existantes
        existing_data = pd.read_csv(csv_filename)
        # Concaténer les anciennes et nouvelles données
        updated_df = pd.concat([existing_data, new_df], ignore_index=True)
        # Supprimer les doublons (si nécessaire)
        updated_df.drop_duplicates(subset=['Nom'], keep='last', inplace=True)
    else:
        # Si le fichier n'existe pas, utiliser les nouvelles données
        updated_df = new_df
    
    # Sauvegarder le fichier CSV mis à jour
    updated_df.to_csv(csv_filename, index=False, encoding='utf-8')
    print(f"CSV file updated and saved to {csv_filename}")
    
    # Sauvegarder le fichier Excel mis à jour
    updated_df.to_excel(excel_filename, index=False)
    print(f"Excel file updated and saved to {excel_filename}")
    
    # Générer le fichier PDF mis à jour
    generate_pdf(updated_df, pdf_filename)
    print(f"PDF file updated and saved to {pdf_filename}")

# Fonction pour générer un PDF à partir d'un DataFrame
def generate_pdf(df, pdf_filename):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=10)
    
    # Ajouter les en-têtes
    headers = df.columns
    for header in headers:
        pdf.cell(40, 10, str(header), border=1)
    pdf.ln()
    
    # Ajouter les données
    for index, row in df.iterrows():
        for item in row:
            pdf.cell(40, 10, str(item), border=1)
        pdf.ln()
    
    # Sauvegarder le PDF
    pdf.output(pdf_filename)

# Fonction pour parcourir toutes les pages de A à Z
def extract_all_data(base_url):
    # Parcourir chaque lettre de l'alphabet
    for letter in 'ABCDEFGHIJKLMNOPQRSTUVWXYZ':
        page_number = 1
        while True:
            url = f"{base_url}/fr/activite/alpha/{letter}?page={page_number}"
            print(f"Extracting data from {url}")
            
            entreprises = extract_page_data(url)
            if not entreprises:
                print(f"No more data for letter {letter}. Moving to next letter.")
                break  # Passer à la lettre suivante si aucune entreprise n'est trouvée
            
            # Mettre à jour les fichiers CSV, Excel et PDF avec les nouvelles données
            update_files(entreprises)
            page_number += 1

# URL de base du site
base_url = 'https://www.doualazoom.com'

# Extraire toutes les données
extract_all_data(base_url)

print("Data extraction and file updates complete.")"""





























"""import requests
from bs4 import BeautifulSoup
import pandas as pd
import os

# Fonction pour extraire les informations d'une page
def extract_page_data(url):
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')
    
    entreprises = []
    
    # Trouver tous les éléments de la liste des entreprises
    for entreprise in soup.select('.div_categorielist_title_container'):
        nom_element = entreprise.select_one('.div_list_nomentreprise a')
        details_element = entreprise.select_one('.div_list_detailentreprise')
        
        if nom_element and details_element:
            nom = nom_element.text.strip()
            
            # Extraire le téléphone
            telephone_element = details_element.find("span", dir="ltr")
            telephone = telephone_element.text.strip() if telephone_element else ""
            
            # Extraire l'email
            email_element = details_element.select_one('a[href^="mailto:"]')
            email = email_element['href'].replace('mailto:', '').strip() if email_element else ""
            
            # Extraire l'adresse
            adresse_parts = [part.text.strip() for part in details_element.select('a') if "ville" in part["href"] or "alpha" in part["href"]]
            adresse = ", ".join(adresse_parts) if adresse_parts else ""
            
            entreprises.append({
                'Nom': nom,
                'Téléphone': telephone,
                'Email': email,
                'Adresse': adresse
            })
    
    return entreprises

# Fonction pour mettre à jour le fichier CSV
def update_csv(new_data, filename='entreprises_yaoundezoom.csv'):
    # Convertir les nouvelles données en DataFrame
    new_df = pd.DataFrame(new_data)
    
    # Vérifier si le fichier existe déjà
    if os.path.exists(filename):
        # Charger les données existantes
        existing_data = pd.read_csv(filename)
        # Concaténer les anciennes et nouvelles données
        updated_df = pd.concat([existing_data, new_df], ignore_index=True)
        # Supprimer les doublons (si nécessaire)
        updated_df.drop_duplicates(subset=['Nom'], keep='last', inplace=True)
    else:
        # Si le fichier n'existe pas, utiliser les nouvelles données
        updated_df = new_df
    
    # Sauvegarder le fichier mis à jour
    updated_df.to_csv(filename, index=False, encoding='utf-8')
    print(f"Data updated and saved to {filename}")

# Fonction pour parcourir toutes les pages de A à Z
def extract_all_data(base_url):
    # Parcourir chaque lettre de l'alphabet
    for letter in 'ABCDEFGHIJKLMNOPQRSTUVWXYZ':
        page_number = 1
        while True:
            url = f"{base_url}/fr/activite/alpha/{letter}?page={page_number}"
            print(f"Extracting data from {url}")
            
            entreprises = extract_page_data(url)
            if not entreprises:
                print(f"No more data for letter {letter}. Moving to next letter.")
                break  # Passer à la lettre suivante si aucune entreprise n'est trouvée
            
            # Mettre à jour le fichier CSV avec les nouvelles données
            update_csv(entreprises)
            page_number += 1

# URL de base du site
base_url = 'https://www.yaoundezoom.com'

# Extraire toutes les données
extract_all_data(base_url)

print("Data extraction and CSV update complete.")"""





























"""import requests
from bs4 import BeautifulSoup
import pandas as pd
import os

# Fonction pour extraire les informations d'une page
def extract_page_data(url):
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')
    
    entreprises = []
    
    # Trouver tous les éléments de la liste des entreprises
    for entreprise in soup.select('.div_categorielist_title_container'):
        nom_element = entreprise.select_one('.div_list_nomentreprise a')
        details_element = entreprise.select_one('.div_list_detailentreprise')
        
        if nom_element and details_element:
            nom = nom_element.text.strip()
            
            # Extraire le téléphone
            telephone_element = details_element.find("span", dir="ltr")
            telephone = telephone_element.text.strip() if telephone_element else ""
            
            # Extraire l'email
            email_element = details_element.select_one('a[href^="mailto:"]')
            email = email_element['href'].replace('mailto:', '').strip() if email_element else ""
            
            # Extraire l'adresse
            adresse_parts = [part.text.strip() for part in details_element.select('a') if "ville" in part["href"] or "alpha" in part["href"]]
            adresse = ", ".join(adresse_parts) if adresse_parts else ""
            
            entreprises.append({
                'Nom': nom,
                'Téléphone': telephone,
                'Email': email,
                'Adresse': adresse
            })
    
    return entreprises

# Fonction pour parcourir toutes les pages de A à Z
def extract_all_data(base_url):
    all_entreprises = []
    
    # Parcourir chaque lettre de l'alphabet
    for letter in 'ABCDEFGHIJKLMNOPQRSTUVWXYZ':
        page_number = 1
        while True:
            url = f"{base_url}/fr/activite/alpha/{letter}?page={page_number}"
            print(f"Extracting data from {url}")
            
            entreprises = extract_page_data(url)
            print(f"Found {len(entreprises)} entreprises.")
            print(entreprises)
            if not entreprises:
                print(f"No more data for letter {letter}. Moving to next letter.")
                break  # Passer à la lettre suivante si aucune entreprise n'est trouvée
            
            all_entreprises.extend(entreprises)
            page_number += 1
    
    return all_entreprises

# Fonction pour mettre à jour le fichier CSV
def update_csv(new_data, filename='entreprises_doualazoom.csv'):
    # Vérifier si le fichier existe déjà
    if os.path.exists(filename):
        # Charger les données existantes
        existing_data = pd.read_csv(filename)
        # Convertir les nouvelles données en DataFrame
        new_df = pd.DataFrame(new_data)
        # Concaténer les anciennes et nouvelles données
        updated_df = pd.concat([existing_data, new_df], ignore_index=True)
        # Supprimer les doublons (si nécessaire)
        updated_df.drop_duplicates(subset=['Nom'], keep='last', inplace=True)
    else:
        # Si le fichier n'existe pas, créer un nouveau DataFrame
        updated_df = pd.DataFrame(new_data)
    
    # Sauvegarder le fichier mis à jour
    updated_df.to_csv(filename, index=False, encoding='utf-8')
    print(f"Data updated and saved to {filename}")

# URL de base du site
base_url = 'https://www.doualazoom.com'

# Extraire toutes les données
all_entreprises = extract_all_data(base_url)

# Mettre à jour le fichier CSV
update_csv(all_entreprises)

print("Data extraction and CSV update complete.")"""    
























"""import requests
from bs4 import BeautifulSoup
import pandas as pd

# Fonction pour extraire les informations d'une page
def extract_page_data(url):
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')
    
    entreprises = []
    
    # Trouver tous les éléments de la liste des entreprises
    for entreprise in soup.select('.div_categorielist_title_container'):
        nom_element = entreprise.select_one('.div_list_nomentreprise a')
        details_element = entreprise.select_one('.div_list_detailentreprise')
        
        if nom_element and details_element:
            nom = nom_element.text.strip()
            
            # Extraire le téléphone
            telephone_element = details_element.find("span", dir="ltr")
            telephone = telephone_element.text.strip() if telephone_element else ""
            
            # Extraire l'email
            email_element = details_element.select_one('a[href^="mailto:"]')
            email = email_element['href'].replace('mailto:', '').strip() if email_element else ""
            
            # Extraire l'adresse
            adresse_parts = [part.text.strip() for part in details_element.select('a') if "ville" in part["href"] or "alpha" in part["href"]]
            adresse = ", ".join(adresse_parts) if adresse_parts else ""
            
            entreprises.append({
                'Nom': nom,
                'Téléphone': telephone,
                'Email': email,
                'Adresse': adresse
            })
    
    return entreprises

# Fonction pour parcourir toutes les pages de A à Z
def extract_all_data(base_url):
    all_entreprises = []
    
    # Parcourir chaque lettre de l'alphabet
    for letter in 'ABCDEFGHIJKLMNOPQRSTUVWXYZ':
        page_number = 1
        while True:
            url = f"{base_url}/fr/activite/alpha/{letter}?page={page_number}"
            print(f"Extracting data from {url}")
            
            entreprises = extract_page_data(url)
            print(f"Found {len(entreprises)} entreprises.")
            print(entreprises)
            if not entreprises:
                print(f"No more data for letter {letter}. Moving to next letter.")
                break  # Passer à la lettre suivante si aucune entreprise n'est trouvée
            
            all_entreprises.extend(entreprises)
            page_number += 1
    
    return all_entreprises

# URL de base du site
base_url = 'https://www.doualazoom.com'

# Extraire toutes les données
all_entreprises = extract_all_data(base_url)

# Convertir en DataFrame et exporter en CSV
df = pd.DataFrame(all_entreprises)
df.to_csv('entreprises_doualazoom.csv', index=False, encoding='utf-8')

print("Data extraction complete. Data saved to 'entreprises_doualazoom.csv'.")"""

"""import requests
from bs4 import BeautifulSoup
import pandas as pd
import os
from fpdf import FPDF  # Bibliothèque pour générer des PDFs

# Fonction pour extraire les informations d'une page
def extract_page_data(url):
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')
    
    entreprises = []
    
    # Trouver tous les éléments de la liste des entreprises
    for entreprise in soup.select('.div_categorielist_title_container'):
        nom_element = entreprise.select_one('.div_list_nomentreprise a')
        details_element = entreprise.select_one('.div_list_detailentreprise')
        
        if nom_element and details_element:
            nom = nom_element.text.strip()
            
            # Extraire le téléphone
            telephone_element = details_element.find("span", dir="ltr")
            telephone = telephone_element.text.strip() if telephone_element else ""
            
            # Extraire l'email
            email_element = details_element.select_one('a[href^="mailto:"]')
            email = email_element['href'].replace('mailto:', '').strip() if email_element else ""
            
            # Extraire l'adresse
            adresse_parts = [part.text.strip() for part in details_element.select('a') if "ville" in part["href"] or "alpha" in part["href"]]
            adresse = ", ".join(adresse_parts) if adresse_parts else ""
            
            entreprises.append({
                'Nom': nom,
                'Téléphone': telephone,
                'Email': email,
                'Adresse': adresse
            })
    
    return entreprises

# Fonction pour mettre à jour le fichier CSV, Excel et PDF
def update_files(new_data, csv_filename='entreprises_doualazoom.csv', excel_filename='entreprises_doualazoom.xlsx', pdf_filename='entreprises_doualazoom.pdf'):
    # Convertir les nouvelles données en DataFrame
    new_df = pd.DataFrame(new_data)
    
    # Vérifier si le fichier CSV existe déjà
    if os.path.exists(csv_filename):
        # Charger les données existantes
        existing_data = pd.read_csv(csv_filename)
        # Concaténer les anciennes et nouvelles données
        updated_df = pd.concat([existing_data, new_df], ignore_index=True)
        # Supprimer les doublons (si nécessaire)
        updated_df.drop_duplicates(subset=['Nom'], keep='last', inplace=True)
    else:
        # Si le fichier n'existe pas, utiliser les nouvelles données
        updated_df = new_df
    
    # Sauvegarder le fichier CSV mis à jour
    updated_df.to_csv(csv_filename, index=False, encoding='utf-8')
    print(f"CSV file updated and saved to {csv_filename}")
    
    # Sauvegarder le fichier Excel mis à jour
    updated_df.to_excel(excel_filename, index=False)
    print(f"Excel file updated and saved to {excel_filename}")
    
    # Générer le fichier PDF mis à jour
    generate_pdf(updated_df, pdf_filename)
    print(f"PDF file updated and saved to {pdf_filename}")

# Fonction pour générer un PDF à partir d'un DataFrame
def generate_pdf(df, pdf_filename):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=10)
    
    # Ajouter les en-têtes
    headers = df.columns
    for header in headers:
        pdf.cell(40, 10, header.encode('latin-1', 'replace').decode('latin-1'), border=1)
    pdf.ln()
    
    # Ajouter les données
    for index, row in df.iterrows():
        for item in row:
            # Encoder les données en latin-1 et remplacer les caractères non supportés
            cell_value = str(item).encode('latin-1', 'replace').decode('latin-1')
            pdf.cell(40, 10, cell_value, border=1)
        pdf.ln()
    
    # Sauvegarder le PDF
    pdf.output(pdf_filename)

# Fonction pour parcourir toutes les pages de A à Z
def extract_all_data(base_url):
    # Parcourir chaque lettre de l'alphabet
    for letter in 'ABCDEFGHIJKLMNOPQRSTUVWXYZ':
        page_number = 1
        while True:
            url = f"{base_url}/fr/activite/alpha/{letter}?page={page_number}"
            print(f"Extracting data from {url}")
            
            entreprises = extract_page_data(url)
            if not entreprises:
                print(f"No more data for letter {letter}. Moving to next letter.")
                break  # Passer à la lettre suivante si aucune entreprise n'est trouvée
            
            # Mettre à jour les fichiers CSV, Excel et PDF avec les nouvelles données
            update_files(entreprises)
            page_number += 1

# URL de base du site
base_url = 'https://www.doualazoom.com'

# Extraire toutes les données
extract_all_data(base_url)

print("Data extraction and file updates complete.")"""
